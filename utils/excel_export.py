"""
Professional Excel report builder for Phishing Simulation campaigns.
Produces a multi-sheet .xlsx workbook with cover page, executive summary,
SBU breakdown, and full target-level detail.
"""

from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

# ── Palette ────────────────────────────────────────────────────────────────
C_NAVY       = "0D1B2A"   # cover background / primary header
C_BLUE       = "1D4ED8"   # accent blue
C_BLUE_MID   = "2563EB"   # table headers
C_BLUE_LIGHT = "DBEAFE"   # column sub-headers / alt rows
C_WHITE      = "FFFFFF"
C_OFFWHITE   = "F8FAFC"
C_LIGHT_GRAY = "E2E8F0"
C_MID_GRAY   = "94A3B8"
C_DARK_TEXT  = "0F172A"
C_GREEN_BG   = "DCFCE7"
C_GREEN_FG   = "15803D"
C_AMBER_BG   = "FEF3C7"
C_AMBER_FG   = "92400E"
C_RED_BG     = "FEE2E2"
C_RED_FG     = "B91C1C"

# ── Helper style factories ─────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=11, color=C_DARK_TEXT, italic=False, name="Calibri"):
    return Font(bold=bold, size=size, color=color, italic=italic, name=name)

def _border_thin():
    s = Side(style="thin", color=C_LIGHT_GRAY)
    return Border(left=s, right=s, top=s, bottom=s)

def _border_medium():
    s = Side(style="medium", color=C_BLUE_MID)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _col(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


_WKHTMLTOIMAGE = r'C:\Program Files\wkhtmltopdf\bin\wkhtmltoimage.exe'


def _try_render_email(html_content):
    """Render HTML email to PNG bytes using imgkit (wkhtmltoimage).
    Returns BytesIO on success, or None if the tool is not installed."""
    if not html_content:
        return None
    try:
        import imgkit
        import os
        config = imgkit.config(wkhtmltoimage=_WKHTMLTOIMAGE) if os.path.exists(_WKHTMLTOIMAGE) else None
        options = {
            'format':             'png',
            'width':              '800',
            'quiet':              '',
            'disable-javascript': '',
        }
        data = imgkit.from_string(html_content, False, options=options, config=config)
        return BytesIO(data)
    except Exception:
        pass
    return None


# ── Cover Sheet ────────────────────────────────────────────────────────────

def _build_cover(ws, campaign, stats):
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False

    # ── background stripe rows 1-6
    for row in range(1, 20):
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.fill = _fill(C_NAVY)

    # Row heights
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 60
    ws.row_dimensions[3].height = 36
    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 14

    # ── Main title  (B2:J2)
    ws.merge_cells("B2:J2")
    t = ws["B2"]
    t.value = "PHISHING SIMULATION"
    t.font = Font(name="Calibri", bold=True, size=38, color=C_WHITE)
    t.alignment = _align("left", "center")

    # ── Sub-title / campaign name (B3:J3)
    ws.merge_cells("B3:J3")
    s = ws["B3"]
    s.value = campaign.name.upper()
    s.font = Font(name="Calibri", bold=False, size=20, color="93C5FD")
    s.alignment = _align("left", "center")

    # ── Tag line (B4:J4)
    ws.merge_cells("B4:J4")
    tl = ws["B4"]
    tl.value = "Security Awareness Campaign Report"
    tl.font = Font(name="Calibri", italic=True, size=12, color=C_MID_GRAY)
    tl.alignment = _align("left", "center")

    # ── Divider row 5: blue stripe
    for col in range(1, 12):
        ws.cell(row=5, column=col).fill = _fill(C_BLUE)

    # ── White body section from row 6
    for row in range(6, 60):
        for col in range(1, 12):
            ws.cell(row=row, column=col).fill = _fill(C_WHITE)

    # --- Column widths
    for c, w in [("A", 3), ("B", 28), ("C", 40), ("D", 5),
                 ("E", 22), ("F", 22), ("G", 22), ("H", 22), ("I", 8), ("J", 8)]:
        _col(ws, c, w)

    # ── Campaign Details table (rows 7-17)
    ws.row_dimensions[6].height = 16
    ws.row_dimensions[7].height = 18

    # Section heading
    ws.merge_cells("B7:C7")
    h = ws["B7"]
    h.value = "CAMPAIGN DETAILS"
    h.font = Font(name="Calibri", bold=True, size=10, color=C_MID_GRAY)
    h.alignment = _align("left", "center")

    def _detail_row(row, label, value):
        ws.row_dimensions[row].height = 22
        lbl = ws.cell(row=row, column=2, value=label)
        lbl.font = Font(name="Calibri", bold=True, size=10, color=C_MID_GRAY)
        lbl.alignment = _align("left", "center")
        val = ws.cell(row=row, column=3, value=value or "—")
        val.font = Font(name="Calibri", size=10, color=C_DARK_TEXT)
        val.alignment = _align("left", "center")
        # Bottom border only
        bd = Side(style="thin", color=C_LIGHT_GRAY)
        for cell in [lbl, val]:
            cell.border = Border(bottom=bd)

    _detail_row(8,  "Campaign Name",  campaign.name)
    _detail_row(9,  "Description",    campaign.description or "—")
    _detail_row(10, "Campaign Type",  (campaign.campaign_type or "—").replace("_", " ").title())
    _detail_row(11, "Email Subject",  campaign.subject or "—")
    _detail_row(12, "Status",         campaign.status.title() if campaign.status else "—")
    _detail_row(13, "Created",        campaign.created_at.strftime("%d %b %Y %H:%M") if campaign.created_at else "—")
    _detail_row(14, "Started",        campaign.started_at.strftime("%d %b %Y %H:%M") if campaign.started_at else "—")
    _detail_row(15, "Finished",       campaign.finished_at.strftime("%d %b %Y %H:%M") if campaign.finished_at else "—")

    # ── KPI boxes (rows 18-27): 5 boxes across columns E-J
    ws.row_dimensions[18].height = 16
    ws.merge_cells("B18:J18")
    kpi_head = ws["B18"]
    kpi_head.value = "CAMPAIGN RESULTS AT A GLANCE"
    kpi_head.font = Font(name="Calibri", bold=True, size=10, color=C_MID_GRAY)
    kpi_head.alignment = _align("left", "bottom")

    total    = stats["total"]
    opened   = stats["opened"]
    clicked  = stats["clicked"]
    reported = stats["reported"]
    comp     = stats["compromised"]
    o_rate   = stats["open_rate"]
    c_rate   = stats["click_rate"]
    r_rate   = stats["report_rate"]
    comp_rate= stats["compromise_rate"]

    kpis = [
        ("Total Targeted",    total,    f"{total} users",      C_BLUE_MID,  C_WHITE),
        ("Opened",            opened,   f"{o_rate}% open rate", "0369A1",   C_WHITE),
        ("Clicked Link",      clicked,  f"{c_rate}% click rate", "DC2626" if c_rate > 15 else "92400E" if c_rate > 5 else "15803D", C_WHITE),
        ("Reported",          reported, f"{r_rate}% report rate", "7C3AED",  C_WHITE),
        ("Compromised",       comp,     f"{comp_rate}% rate",   "B91C1C" if comp > 0 else "374151", C_WHITE),
    ]

    # Columns B C D E F (with gaps), each box spans 2 cols
    box_cols = [2, 4, 6, 8, 10]   # starting column numbers (1-based)
    box_span = 1                   # each box is 1 col wide for simplicity
    # We'll use a 2-row box: label in row 19, value in row 20, subtext in row 21
    for i, (label, value, sub, bg, fg) in enumerate(kpis):
        col = box_cols[i]
        for r in range(19, 24):
            ws.row_dimensions[r].height = 20 if r in (20,) else 16
        # Number cell
        num_cell = ws.cell(row=20, column=col, value=value)
        num_cell.font = Font(name="Calibri", bold=True, size=28, color=bg)
        num_cell.alignment = _align("left", "center")

        lbl_cell = ws.cell(row=19, column=col, value=label)
        lbl_cell.font = Font(name="Calibri", bold=True, size=9, color=C_MID_GRAY)
        lbl_cell.alignment = _align("left", "center")

        sub_cell = ws.cell(row=21, column=col, value=sub)
        sub_cell.font = Font(name="Calibri", size=9, color=bg)
        sub_cell.alignment = _align("left", "center")

    # ── Report generated date (bottom)
    ws.row_dimensions[27].height = 18
    gen = ws.cell(row=27, column=2,
                  value=f"Report generated: {datetime.utcnow().strftime('%d %b %Y at %H:%M UTC')}")
    gen.font = Font(name="Calibri", italic=True, size=9, color=C_MID_GRAY)


# ── Executive Summary Sheet ────────────────────────────────────────────────

def _build_summary(ws, campaign, stats, results_data):
    ws.title = "Executive Summary"
    ws.sheet_view.showGridLines = False

    for c, w in [("A", 3), ("B", 28), ("C", 18), ("D", 18),
                 ("E", 18), ("F", 18), ("G", 18), ("H", 13), ("I", 13), ("J", 3)]:
        _col(ws, c, w)

    # Header band — merge starts at B1 (the cell we write to); A1 filled separately
    ws.cell(row=1, column=1).fill = _fill(C_BLUE_MID)
    ws.merge_cells("B1:J1")
    for col in range(2, 11):
        ws.cell(row=1, column=col).fill = _fill(C_BLUE_MID)
    ws.row_dimensions[1].height = 48

    title = ws["B1"]
    title.value = "Executive Summary"
    title.font = Font(name="Calibri", bold=True, size=20, color=C_WHITE)
    title.alignment = _align("left", "center")

    ws.row_dimensions[2].height = 8

    # ── Key Metrics table ──────────────────────────────────────────────────
    def _section(row, text):
        ws.merge_cells(f"B{row}:G{row}")
        c = ws[f"B{row}"]
        c.value = text
        c.font = Font(name="Calibri", bold=True, size=11, color=C_BLUE_MID)
        c.fill = _fill(C_BLUE_LIGHT)
        c.alignment = _align("left", "center")
        ws.row_dimensions[row].height = 24

    def _metric_row(row, label, raw_val, rate_val, rate_label, fill_color=C_WHITE):
        ws.row_dimensions[row].height = 22
        cells_data = [
            ("B", label, True, 10, C_DARK_TEXT),
            ("C", raw_val, False, 14, C_BLUE_MID),
            ("D", rate_val, False, 14, C_RED_FG if isinstance(rate_val, (int, float)) and rate_val > 15 else C_GREEN_FG),
            ("E", rate_label, False, 9, C_MID_GRAY),
        ]
        for col_letter, value, bold, size, color in cells_data:
            cell = ws[f"{col_letter}{row}"]
            cell.value = value
            cell.font = Font(name="Calibri", bold=bold, size=size, color=color)
            cell.fill = _fill(fill_color if col_letter == "B" else C_WHITE)
            cell.alignment = _align("left" if col_letter in ("B", "E") else "center", "center")
            cell.border = _border_thin()

    _section(3, "OVERALL CAMPAIGN PERFORMANCE")

    # Column headers
    ws.row_dimensions[4].height = 20
    for col_l, text in [("B", "Metric"), ("C", "Count"), ("D", "Rate (%)"), ("E", "Benchmark")]:
        hdr = ws[f"{col_l}4"]
        hdr.value = text
        hdr.font = Font(name="Calibri", bold=True, size=10, color=C_WHITE)
        hdr.fill = _fill(C_NAVY)
        hdr.alignment = _align("center", "center")
        hdr.border = _border_thin()

    t = stats["total"]
    rows_data = [
        ("Emails Targeted",        t,                                          None,                               ""),
        ("Emails Opened",          stats["opened"],                            stats["open_rate"],                 "> 30% is common"),
        ("Links Clicked",          stats["clicked"],                           stats["click_rate"],                "< 5% is good"),
        ("Questionnaire Finished", stats.get("questionnaire_finished", 0),     stats.get("questionnaire_rate", 0.0), "> 50% is good"),
        ("Reported to Security",   stats["reported"],                          stats["report_rate"],               "> 10% is healthy"),
    ]
    fill_alt = [C_WHITE, C_OFFWHITE]
    for i, (lbl, cnt, rate, bench) in enumerate(rows_data):
        r = 5 + i
        ws.row_dimensions[r].height = 22
        fill = fill_alt[i % 2]
        ws[f"B{r}"].value = lbl
        ws[f"B{r}"].font = Font(name="Calibri", size=10, color=C_DARK_TEXT)
        ws[f"B{r}"].fill = _fill(fill)
        ws[f"B{r}"].alignment = _align("left", "center")
        ws[f"B{r}"].border = _border_thin()

        ws[f"C{r}"].value = cnt
        ws[f"C{r}"].font = Font(name="Calibri", bold=True, size=12, color=C_BLUE_MID)
        ws[f"C{r}"].fill = _fill(fill)
        ws[f"C{r}"].alignment = _align("center", "center")
        ws[f"C{r}"].border = _border_thin()

        if rate is not None:
            rate_color = C_RED_FG  if (lbl == "Links Clicked" and rate > 15) \
                         else C_AMBER_FG if (lbl == "Links Clicked" and rate > 5) \
                         else C_GREEN_FG
            rate_fill  = C_RED_BG  if (lbl == "Links Clicked" and rate > 15) \
                         else C_AMBER_BG if (lbl == "Links Clicked" and rate > 5) \
                         else C_GREEN_BG
            ws[f"D{r}"].value = f"{rate}%"
            ws[f"D{r}"].font = Font(name="Calibri", bold=True, size=12, color=rate_color)
            ws[f"D{r}"].fill = _fill(rate_fill)
        else:
            ws[f"D{r}"].value = "—"
            ws[f"D{r}"].font = Font(name="Calibri", size=10, color=C_MID_GRAY)
            ws[f"D{r}"].fill = _fill(fill)
        ws[f"D{r}"].alignment = _align("center", "center")
        ws[f"D{r}"].border = _border_thin()

        ws[f"E{r}"].value = bench
        ws[f"E{r}"].font = Font(name="Calibri", italic=True, size=9, color=C_MID_GRAY)
        ws[f"E{r}"].fill = _fill(fill)
        ws[f"E{r}"].alignment = _align("left", "center")
        ws[f"E{r}"].border = _border_thin()

    # Campaign metadata block
    _section(12, "CAMPAIGN INFORMATION")
    meta = [
        ("Campaign Name",    campaign.name),
        ("Description",      campaign.description or "—"),
        ("Type",             (campaign.campaign_type or "—").replace("_", " ").title()),
        ("Email Subject",    campaign.subject or "—"),
        ("Status",           campaign.status.title() if campaign.status else "—"),
        ("Created",          campaign.created_at.strftime("%d %b %Y") if campaign.created_at else "—"),
        ("Started",          campaign.started_at.strftime("%d %b %Y %H:%M") if campaign.started_at else "—"),
        ("Finished",         campaign.finished_at.strftime("%d %b %Y %H:%M") if campaign.finished_at else "—"),
    ]
    for i, (lbl, val) in enumerate(meta):
        r = 13 + i
        ws.row_dimensions[r].height = 20
        fill = fill_alt[i % 2]
        lc = ws[f"B{r}"]
        lc.value = lbl
        lc.font = Font(name="Calibri", bold=True, size=10, color=C_DARK_TEXT)
        lc.fill = _fill(fill)
        lc.alignment = _align("left", "center")
        lc.border = _border_thin()
        vc = ws[f"C{r}"]
        ws.merge_cells(f"C{r}:G{r}")
        vc.value = val
        vc.font = Font(name="Calibri", size=10, color=C_DARK_TEXT)
        vc.fill = _fill(fill)
        vc.alignment = _align("left", "center", wrap=True)
        vc.border = _border_thin()

    # ── Dynamic analysis sections (row 22 onward) ─────────────────────────
    cur = 22  # row 21 = visual spacer gap

    def _hdr_and_total(row, title, count, end_col="G"):
        """Write navy heading row + total subtitle. Returns next available row."""
        ws.row_dimensions[row].height = 28
        ws.merge_cells(f"B{row}:{end_col}{row}")
        h = ws[f"B{row}"]
        h.value = title
        h.font = Font(name="Calibri", bold=True, size=12, color=C_WHITE)
        h.fill = _fill(C_NAVY)
        h.alignment = _align("left", "center")
        ws.row_dimensions[row + 1].height = 22
        ws.merge_cells(f"B{row+1}:{end_col}{row+1}")
        tc = ws[f"B{row+1}"]
        tc.value = f"Total: {count}"
        tc.font = Font(name="Calibri", bold=True, size=11, color=C_BLUE_MID)
        tc.fill = _fill(C_BLUE_LIGHT)
        tc.alignment = _align("left", "center")
        return row + 2

    def _col_hdrs(row, cols):
        """Write column header row. Returns next row."""
        ws.row_dimensions[row].height = 22
        for col_l, text in cols:
            c = ws[f"{col_l}{row}"]
            c.value = text
            c.font = Font(name="Calibri", bold=True, size=10, color=C_WHITE)
            c.fill = _fill(C_BLUE_MID)
            c.alignment = _align("center", "center")
            c.border = _border_thin()
        return row + 1

    def _data_r(row, vals, base_fill=C_WHITE):
        """Write a data row. Each val: (col_l, value, bold, fg, align_h, cell_fill_or_None)."""
        ws.row_dimensions[row].height = 20
        for col_l, val, bold, fg, align_h, cell_fill in vals:
            c = ws[f"{col_l}{row}"]
            c.value = val
            c.font = Font(name="Calibri", bold=bold, size=10, color=fg)
            c.fill = _fill(cell_fill if cell_fill else base_fill)
            c.alignment = _align(align_h, "center")
            c.border = _border_thin()

    def _empty_note(row, text):
        ws.row_dimensions[row].height = 20
        ws.merge_cells(f"B{row}:G{row}")
        c = ws[f"B{row}"]
        c.value = text
        c.font = Font(name="Calibri", italic=True, size=10, color=C_MID_GRAY)
        c.alignment = _align("left", "center")

    # ── SBU Breakdown ──────────────────────────────────────────────────────
    sbu_list = stats.get("sbu_list", [])
    ws.row_dimensions[cur].height = 8
    cur += 1
    cur = _hdr_and_total(cur, "SBU BREAKDOWN", len(sbu_list), end_col="I")
    cur = _col_hdrs(cur, [
        ("B", "Department / SBU"), ("C", "Targeted"), ("D", "Opened"),
        ("E", "Clicked"),          ("F", "Click Rate"), ("G", "Reported"),
        ("H", "Report Rate"),      ("I", "Compromised"),
    ])
    for i, sbu in enumerate(sbu_list):
        cr = sbu["click_rate"]
        ck_fill  = C_RED_BG if cr > 15 else C_AMBER_BG if cr > 5 else C_GREEN_BG
        ck_color = C_RED_FG if cr > 15 else C_AMBER_FG if cr > 5 else C_GREEN_FG
        cm_color = C_RED_FG if sbu["compromised"] > 0 else C_DARK_TEXT
        cm_fill  = C_RED_BG if sbu["compromised"] > 0 else None
        base = fill_alt[i % 2]
        _data_r(cur, [
            ("B", sbu["sbu"],              False, C_DARK_TEXT, "left",   None),
            ("C", sbu["targeted"],          True,  C_BLUE_MID,  "center", None),
            ("D", sbu["opened"],            False, C_DARK_TEXT, "center", None),
            ("E", sbu["clicked"],           True,  ck_color,    "center", ck_fill),
            ("F", f'{sbu["click_rate"]}%',  True,  ck_color,    "center", ck_fill),
            ("G", sbu["reported"],          False, C_DARK_TEXT, "center", None),
            ("H", f'{sbu["report_rate"]}%', False, C_DARK_TEXT, "center", None),
            ("I", sbu["compromised"],       True,  cm_color,    "center", cm_fill),
        ], base)
        cur += 1
    if sbu_list:
        cr = stats["click_rate"]
        ck_f = C_RED_BG if cr > 15 else C_AMBER_BG if cr > 5 else C_GREEN_BG
        ck_c = C_RED_FG if cr > 15 else C_AMBER_FG if cr > 5 else C_GREEN_FG
        _data_r(cur, [
            ("B", "TOTAL",                    True,  C_WHITE, "left",   C_NAVY),
            ("C", stats["total"],             True,  C_WHITE, "center", C_NAVY),
            ("D", stats["opened"],            True,  C_WHITE, "center", C_NAVY),
            ("E", stats["clicked"],           True,  ck_c,    "center", ck_f),
            ("F", f'{cr}%',                   True,  ck_c,    "center", ck_f),
            ("G", stats["reported"],          True,  C_WHITE, "center", C_NAVY),
            ("H", f'{stats["report_rate"]}%', True,  C_WHITE, "center", C_NAVY),
            ("I", stats["compromised"],       True,
             C_WHITE if stats["compromised"] == 0 else C_RED_FG,
             "center",
             C_NAVY  if stats["compromised"] == 0 else C_RED_BG),
        ], C_NAVY)
        cur += 1

    # ── Employees Who Clicked ──────────────────────────────────────────────
    clickers = [rd for rd in results_data if rd.get("clicked")]
    ws.row_dimensions[cur].height = 8
    cur += 1
    cur = _hdr_and_total(cur, "EMPLOYEES WHO CLICKED", len(clickers))
    if clickers:
        cur = _col_hdrs(cur, [
            ("B", "Email"), ("C", "First Name"), ("D", "Last Name"),
            ("E", "SBU / Dept"), ("F", "Status"),
        ])
        for i, rd in enumerate(clickers):
            reported = rd.get("reported", False)
            status   = "Recovered"    if reported else "Compromised"
            s_color  = C_GREEN_FG     if reported else C_RED_FG
            s_fill   = C_GREEN_BG     if reported else C_RED_BG
            _data_r(cur, [
                ("B", rd.get("email",      ""), False, C_DARK_TEXT, "left",   None),
                ("C", rd.get("first_name", ""), False, C_DARK_TEXT, "left",   None),
                ("D", rd.get("last_name",  ""), False, C_DARK_TEXT, "left",   None),
                ("E", rd.get("sbu",        ""), False, C_MID_GRAY,  "left",   None),
                ("F", status,                    True,  s_color,     "center", s_fill),
            ], fill_alt[i % 2])
            cur += 1
    else:
        _empty_note(cur, "No employees clicked in this campaign.")
        cur += 1

    # ── Compromised Users ──────────────────────────────────────────────────
    compromised_list = [rd for rd in results_data if rd.get("compromised")]
    ws.row_dimensions[cur].height = 8
    cur += 1
    cur = _hdr_and_total(cur, "COMPROMISED USERS", len(compromised_list))
    if compromised_list:
        cur = _col_hdrs(cur, [
            ("B", "Email"), ("C", "First Name"), ("D", "Last Name"),
            ("E", "SBU / Dept"), ("F", ""),
        ])
        for i, rd in enumerate(compromised_list):
            _data_r(cur, [
                ("B", rd.get("email",      ""), False, C_DARK_TEXT, "left",   None),
                ("C", rd.get("first_name", ""), False, C_DARK_TEXT, "left",   None),
                ("D", rd.get("last_name",  ""), False, C_DARK_TEXT, "left",   None),
                ("E", rd.get("sbu",        ""), False, C_MID_GRAY,  "left",   None),
                ("F", "Compromised",             True,  C_RED_FG,   "center", C_RED_BG),
            ], fill_alt[i % 2])
            cur += 1
    else:
        _empty_note(cur, "No compromised users in this campaign.")
        cur += 1

    # ── Reporters ──────────────────────────────────────────────────────────
    reporters_list = [rd for rd in results_data if rd.get("reported")]
    ws.row_dimensions[cur].height = 8
    cur += 1
    cur = _hdr_and_total(cur, "REPORTERS", len(reporters_list))
    if reporters_list:
        cur = _col_hdrs(cur, [
            ("B", "Email"), ("C", "First Name"), ("D", "Last Name"),
            ("E", "SBU / Dept"), ("F", ""),
        ])
        for i, rd in enumerate(reporters_list):
            _data_r(cur, [
                ("B", rd.get("email",      ""), False, C_DARK_TEXT, "left",   None),
                ("C", rd.get("first_name", ""), False, C_DARK_TEXT, "left",   None),
                ("D", rd.get("last_name",  ""), False, C_DARK_TEXT, "left",   None),
                ("E", rd.get("sbu",        ""), False, C_MID_GRAY,  "left",   None),
                ("F", "Reported",               True,  C_GREEN_FG, "center", C_GREEN_BG),
            ], fill_alt[i % 2])
            cur += 1
    else:
        _empty_note(cur, "No reporters in this campaign.")
        cur += 1


# ── SBU Breakdown Sheet ────────────────────────────────────────────────────

def _build_sbu(ws, stats):
    ws.title = "SBU Breakdown"
    ws.sheet_view.showGridLines = False

    for c, w in [("A", 3), ("B", 28), ("C", 13), ("D", 13), ("E", 14),
                 ("F", 13), ("G", 14), ("H", 13), ("I", 13), ("J", 14), ("K", 3)]:
        _col(ws, c, w)

    # Header band
    ws.cell(row=1, column=1).fill = _fill(C_BLUE_MID)  # A1 — not part of merge
    ws.merge_cells("B1:K1")
    for col in range(2, 12):
        ws.cell(row=1, column=col).fill = _fill(C_BLUE_MID)
    ws.row_dimensions[1].height = 48
    title = ws["B1"]
    title.value = "SBU / Department Breakdown"
    title.font = Font(name="Calibri", bold=True, size=20, color=C_WHITE)
    title.alignment = _align("left", "center")

    ws.row_dimensions[2].height = 8

    # Column headers
    headers = ["Department / SBU", "Targeted", "Opened", "Open Rate",
               "Clicked", "Click Rate", "Reported", "Report Rate",
               "Compromised", "Compromise Rate"]
    header_cols = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K2"]
    col_letters = ["B", "C", "D", "E", "F", "G", "H", "I", "J"]

    ws.row_dimensions[3].height = 30
    for i, (col_l, hdr) in enumerate(zip(col_letters, headers)):
        c = ws[f"{col_l}3"]
        c.value = hdr
        c.font = Font(name="Calibri", bold=True, size=10, color=C_WHITE)
        c.fill = _fill(C_NAVY)
        c.alignment = _align("center", "center", wrap=True)
        c.border = _border_thin()

    sbu_list = stats.get("sbu_list", [])
    fill_alt = [C_WHITE, C_OFFWHITE]

    for i, sbu in enumerate(sbu_list):
        r = 4 + i
        ws.row_dimensions[r].height = 22
        fill = fill_alt[i % 2]

        # Click rate color coding
        cr = sbu["click_rate"]
        click_fill = C_RED_BG if cr > 15 else C_AMBER_BG if cr > 5 else C_GREEN_BG
        click_color = C_RED_FG if cr > 15 else C_AMBER_FG if cr > 5 else C_GREEN_FG

        comp_fill  = C_RED_BG if sbu["compromised"] > 0 else fill
        comp_color = C_RED_FG if sbu["compromised"] > 0 else C_DARK_TEXT

        data = [
            ("B", sbu["sbu"],               False, fill,       C_DARK_TEXT, "left"),
            ("C", sbu["targeted"],           True,  fill,       C_BLUE_MID,  "center"),
            ("D", sbu["opened"],             False, fill,       C_DARK_TEXT, "center"),
            ("E", f'{sbu["open_rate"]}%',    False, fill,       C_DARK_TEXT, "center"),
            ("F", sbu["clicked"],            True,  click_fill, click_color, "center"),
            ("G", f'{sbu["click_rate"]}%',   True,  click_fill, click_color, "center"),
            ("H", sbu["reported"],           False, fill,       C_DARK_TEXT, "center"),
            ("I", f'{sbu["report_rate"]}%',  False, fill,       C_DARK_TEXT, "center"),
            ("J", sbu["compromised"],        True,  comp_fill,  comp_color,  "center"),
        ]
        for col_l, val, bold, bg, fg, align_h in data:
            cell = ws[f"{col_l}{r}"]
            cell.value = val
            cell.font = Font(name="Calibri", bold=bold, size=10, color=fg)
            cell.fill = _fill(bg)
            cell.alignment = _align(align_h, "center")
            cell.border = _border_thin()

    # Totals row
    if sbu_list:
        tr = 4 + len(sbu_list)
        ws.row_dimensions[tr].height = 24
        t = stats["total"]
        cr = stats["click_rate"]
        totals = [
            ("B", "TOTAL",                 True,  C_NAVY, C_WHITE, "left"),
            ("C", t,                        True,  C_NAVY, C_WHITE, "center"),
            ("D", stats["opened"],          True,  C_NAVY, C_WHITE, "center"),
            ("E", f'{stats["open_rate"]}%', True,  C_NAVY, C_WHITE, "center"),
            ("F", stats["clicked"],         True,  C_NAVY, C_WHITE, "center"),
            ("G", f'{cr}%',                 True,  C_NAVY, C_WHITE, "center"),
            ("H", stats["reported"],        True,  C_NAVY, C_WHITE, "center"),
            ("I", f'{stats["report_rate"]}%',True, C_NAVY, C_WHITE, "center"),
            ("J", stats["compromised"],     True,  C_NAVY, C_WHITE, "center"),
        ]
        for col_l, val, bold, bg, fg, align_h in totals:
            cell = ws[f"{col_l}{tr}"]
            cell.value = val
            cell.font = Font(name="Calibri", bold=bold, size=10, color=fg)
            cell.fill = _fill(bg)
            cell.alignment = _align(align_h, "center")
            cell.border = _border_thin()

    # Legend
    legend_row = 4 + len(sbu_list) + 2
    ws.row_dimensions[legend_row].height = 16
    ws.merge_cells(f"B{legend_row}:D{legend_row}")
    lh = ws[f"B{legend_row}"]
    lh.value = "CLICK RATE RISK KEY"
    lh.font = Font(name="Calibri", bold=True, size=9, color=C_MID_GRAY)
    lh.alignment = _align("left", "center")

    legend_items = [
        (C_GREEN_BG, C_GREEN_FG, "≤ 5%  — Low Risk"),
        (C_AMBER_BG, C_AMBER_FG, "5–15%  — Moderate Risk"),
        (C_RED_BG,   C_RED_FG,   "> 15% — High Risk"),
    ]
    lr = legend_row + 1
    for bg, fg, label in legend_items:
        ws.row_dimensions[lr].height = 18
        ws.merge_cells(f"B{lr}:D{lr}")
        c = ws[f"B{lr}"]
        c.value = label
        c.font = Font(name="Calibri", size=9, color=fg)
        c.fill = _fill(bg)
        c.alignment = _align("left", "center")
        c.border = _border_thin()
        lr += 1


# ── Target Detail Sheet ────────────────────────────────────────────────────

def _build_targets(ws, results_data):
    ws.title = "Target Details"
    ws.sheet_view.showGridLines = False

    col_widths = [("A", 3), ("B", 26), ("C", 18), ("D", 18), ("E", 22),
                  ("F", 10), ("G", 10), ("H", 10), ("I", 10), ("J", 10),
                  ("K", 22), ("L", 10), ("M", 3)]
    for c, w in col_widths:
        _col(ws, c, w)

    # Header band
    ws.cell(row=1, column=1).fill = _fill(C_BLUE_MID)  # A1 — not part of merge
    ws.merge_cells("B1:M1")
    for col in range(2, 14):
        ws.cell(row=1, column=col).fill = _fill(C_BLUE_MID)
    ws.row_dimensions[1].height = 48
    title = ws["B1"]
    title.value = "Individual Target Results"
    title.font = Font(name="Calibri", bold=True, size=20, color=C_WHITE)
    title.alignment = _align("left", "center")

    ws.row_dimensions[2].height = 8

    # Column headers
    headers = ["Email", "First Name", "Last Name", "SBU / Department",
               "Opened", "Clicked", "Submitted", "Reported",
               "Compromised", "Clicked At", "Reminders"]
    col_letters = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]

    ws.row_dimensions[3].height = 28
    for col_l, hdr in zip(col_letters, headers):
        c = ws[f"{col_l}3"]
        c.value = hdr
        c.font = Font(name="Calibri", bold=True, size=10, color=C_WHITE)
        c.fill = _fill(C_NAVY)
        c.alignment = _align("center", "center", wrap=True)
        c.border = _border_thin()

    fill_alt = [C_WHITE, C_OFFWHITE]

    for i, row_data in enumerate(results_data):
        r = 4 + i
        ws.row_dimensions[r].height = 20
        fill = fill_alt[i % 2]

        clicked     = row_data.get("clicked", False)
        compromised = row_data.get("compromised", False)
        reported    = row_data.get("reported", False)

        row_fill = C_RED_BG if compromised else C_AMBER_BG if clicked else fill
        bool_yes_fill = _fill(C_GREEN_BG)
        bool_no_fill  = _fill(fill)

        def _yn(val):
            return "Yes" if val else "No"

        def _yn_cell(col_l, val):
            c = ws[f"{col_l}{r}"]
            c.value = _yn(val)
            if val:
                c.font = Font(name="Calibri", bold=True, size=10, color=C_GREEN_FG)
                c.fill = _fill(C_GREEN_BG)
            else:
                c.font = Font(name="Calibri", size=10, color=C_MID_GRAY)
                c.fill = _fill(fill)
            c.alignment = _align("center", "center")
            c.border = _border_thin()

        cells = [
            ("B", row_data.get("email", ""),       False, row_fill, C_DARK_TEXT, "left"),
            ("C", row_data.get("first_name", ""),  False, row_fill, C_DARK_TEXT, "left"),
            ("D", row_data.get("last_name", ""),   False, row_fill, C_DARK_TEXT, "left"),
            ("E", row_data.get("sbu", ""),         False, row_fill, C_DARK_TEXT, "left"),
        ]
        for col_l, val, bold, bg, fg, align_h in cells:
            cell = ws[f"{col_l}{r}"]
            cell.value = val
            cell.font = Font(name="Calibri", bold=bold, size=10, color=fg)
            cell.fill = _fill(bg)
            cell.alignment = _align(align_h, "center")
            cell.border = _border_thin()

        # Boolean columns
        for col_l, key in [("F", "opened"), ("G", "clicked"), ("H", "submitted"),
                           ("I", "reported"), ("J", "compromised")]:
            _yn_cell(col_l, row_data.get(key, False))

        # Clicked at
        ca = ws[f"K{r}"]
        ca.value = row_data.get("clicked_at", "")
        ca.font = Font(name="Calibri", size=9, color=C_MID_GRAY)
        ca.fill = _fill(row_fill)
        ca.alignment = _align("center", "center")
        ca.border = _border_thin()

        # Reminders sent
        rm = ws[f"L{r}"]
        rm.value = row_data.get("reminder_count", 0)
        rm.font = Font(name="Calibri", size=10, color=C_DARK_TEXT)
        rm.fill = _fill(row_fill)
        rm.alignment = _align("center", "center")
        rm.border = _border_thin()

    # Freeze header rows
    ws.freeze_panes = "B4"


# ── Main entry point ───────────────────────────────────────────────────────

def _build_html_sheet(ws, sheet_title, html_content, render_width=900):
    """Render an HTML string to PNG via imgkit and embed it as a full-sheet image."""
    ws.title = sheet_title
    ws.sheet_view.showGridLines = False

    # Header band
    ws.row_dimensions[1].height = 36
    ws.merge_cells("B1:L1")
    hdr = ws["B1"]
    hdr.value = sheet_title.upper()
    hdr.font = Font(name="Calibri", bold=True, size=14, color=C_WHITE)
    hdr.fill = _fill(C_NAVY)
    hdr.alignment = _align("left", "center")
    ws.cell(row=1, column=1).fill = _fill(C_NAVY)

    img_buf = _try_render_email(html_content)
    if img_buf:
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(img_buf)
            # Scale to fit nicely: max width ~900px → ~675pt at 72dpi approx
            max_w = 900
            orig_w, orig_h = img.width, img.height
            scale = min(1.0, max_w / orig_w) if orig_w else 1.0
            img.width  = int(orig_w * scale)
            img.height = int(orig_h * scale)
            img.anchor = "B3"
            ws.add_image(img)
            # set row heights to accommodate the image
            rows_needed = max(30, int(img.height / 15) + 2)
            for rr in range(3, 3 + rows_needed):
                ws.row_dimensions[rr].height = 15
        except Exception as e:
            ws.row_dimensions[3].height = 24
            ws.merge_cells("B3:L3")
            c = ws["B3"]
            c.value = f"Image embed failed: {e}"
            c.font = Font(name="Calibri", italic=True, size=10, color=C_RED_FG)
            c.alignment = _align("left", "center")
    else:
        ws.row_dimensions[3].height = 24
        ws.merge_cells("B3:L3")
        c = ws["B3"]
        c.value = "\u26a0  Preview unavailable — wkhtmltoimage not found or rendering failed."
        c.font = Font(name="Calibri", italic=True, size=10, color=C_MID_GRAY)
        c.alignment = _align("left", "center")

    for col, width in [("A", 2), ("B", 18)] + [(get_column_letter(i), 12) for i in range(3, 13)]:
        ws.column_dimensions[col].width = width


def build_campaign_xlsx(campaign, stats, results_data, email_html=None, landing_page_html=None):
    """
    Build and return a BytesIO containing the full .xlsx report.

    :param campaign:          Campaign ORM object
    :param stats:             Dict from build_stats() — totals + sbu_list
    :param results_data:      List of dicts, one per result row
    :param email_html:        Optional HTML of the campaign phishing email
    :param landing_page_html: Optional HTML of the campaign landing page
    :return:                  BytesIO positioned at 0
    """
    wb = Workbook()
    wb.remove(wb.active)

    ws_cover   = wb.create_sheet("Cover")
    ws_summary = wb.create_sheet("Executive Summary")
    ws_targets = wb.create_sheet("Target Details")

    _build_cover(ws_cover, campaign, stats)
    _build_summary(ws_summary, campaign, stats, results_data)
    _build_targets(ws_targets, results_data)

    if email_html:
        ws_email = wb.create_sheet("Email Preview")
        _build_html_sheet(ws_email, "Campaign Email Preview", email_html)

    if landing_page_html:
        ws_lp = wb.create_sheet("Landing Page")
        _build_html_sheet(ws_lp, "Campaign Landing Page", landing_page_html)

    wb.active = ws_cover

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
